import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

data = json.loads(sys.stdin.read())

wb = Workbook()

# ── colours ──────────────────────────────────────────────────────────
C_DARK   = "0F1623"   # sidebar dark
C_TEAL   = "00B894"   # accent teal
C_LIGHT  = "F0FDF9"   # light teal tint
C_GREY   = "F7FAFC"   # row zebra
C_WHITE  = "FFFFFF"
C_RED    = "E53E3E"
C_TEXT   = "1A1F2E"
C_MUTED  = "718096"

def hdr_fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)
def font(bold=False, color=C_TEXT, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)
def bottom_border():
    s = Side(style="medium", color=C_TEAL)
    return Border(bottom=s)
def top_border():
    s = Side(style="thin", color="CBD5E0")
    return Border(top=s)

UGX = '#,##0;(#,##0);"-"'

def fmt_num(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val if val else 0)
    c.number_format = UGX
    c.alignment = Alignment(horizontal="right")
    return c

def section_header(ws, row, label, bg=C_DARK, fg=C_WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=label)
    c.font = font(bold=True, color=fg, size=10)
    c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = hdr_fill(bg)

def total_row(ws, row, label, formula, bold=True, bg=C_LIGHT, fg=C_TEXT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font = font(bold=bold, color=fg)
    c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal="left", indent=2)
    v = ws.cell(row=row, column=4, value=formula)
    v.font = font(bold=bold, color=fg)
    v.number_format = UGX
    v.fill = hdr_fill(bg)
    v.alignment = Alignment(horizontal="right")
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = bottom_border()
    ws.row_dimensions[row].height = 17

def data_row(ws, row, label, amount, zebra=False):
    bg = C_GREY if zebra else C_WHITE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font = font(color=C_TEXT)
    c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal="left", indent=3)
    v = fmt_num(ws, row, 4, amount)
    v.fill = hdr_fill(bg)
    ws.row_dimensions[row].height = 15

def blank_row(ws, row):
    ws.row_dimensions[row].height = 6

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════════════════════════
#  SHEET 1 — INCOME & EXPENDITURE STATEMENT
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Income & Expenditure"
col_widths(ws1, [28, 4, 4, 18])

r = 1
# Title block
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws1.cell(row=r, column=1, value=data["sacco_name"].upper())
c.font = font(bold=True, color=C_DARK, size=13)
c.alignment = Alignment(horizontal="center")
ws1.row_dimensions[r].height = 22; r += 1

ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws1.cell(row=r, column=1, value="INCOME & EXPENDITURE STATEMENT")
c.font = font(bold=True, color=C_TEAL, size=11)
c.alignment = Alignment(horizontal="center")
ws1.row_dimensions[r].height = 18; r += 1

ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws1.cell(row=r, column=1, value=f"For the period {data['start_date']} to {data['end_date']}")
c.font = font(italic=True, color=C_MUTED, size=9)
c.alignment = Alignment(horizontal="center")
ws1.row_dimensions[r].height = 14; r += 1

blank_row(ws1, r); r += 1

# Column headers
for col, lbl in [(1,"Description"), (4,"Amount (UGX)")]:
    c = ws1.cell(row=r, column=col, value=lbl)
    c.font = font(bold=True, color=C_WHITE)
    c.fill = hdr_fill(C_DARK)
    c.alignment = Alignment(horizontal="left" if col==1 else "right", indent=1)
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws1.row_dimensions[r].height = 16; r += 1

# ── INCOME ──────────────────────────────────────────────────────
section_header(ws1, r, "INCOME", C_TEAL, C_WHITE); r += 1
income_start = r
for i, item in enumerate(data["income_items"]):
    data_row(ws1, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
income_end = r - 1
total_row(ws1, r, "TOTAL INCOME", f"=SUM(D{income_start}:D{income_end})", bg="E6FAF4", fg=C_DARK); 
total_income_row = r; r += 1

blank_row(ws1, r); r += 1

# ── EXPENDITURE ─────────────────────────────────────────────────
section_header(ws1, r, "EXPENDITURE", C_RED, C_WHITE); r += 1
exp_start = r
for i, item in enumerate(data["expenditure_items"]):
    data_row(ws1, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
exp_end = r - 1
total_row(ws1, r, "TOTAL EXPENDITURE", f"=SUM(D{exp_start}:D{exp_end})", bg="FFF5F5", fg=C_RED)
total_exp_row = r; r += 1

blank_row(ws1, r); r += 1

# ── NET SURPLUS / DEFICIT ────────────────────────────────────────
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws1.cell(row=r, column=1, value="NET SURPLUS / (DEFICIT)")
c.font = font(bold=True, color=C_WHITE, size=11)
c.fill = hdr_fill(C_DARK)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
v = ws1.cell(row=r, column=4, value=f"=D{total_income_row}-D{total_exp_row}")
v.font = font(bold=True, color=C_WHITE, size=11)
v.number_format = UGX
v.fill = hdr_fill(C_DARK)
v.alignment = Alignment(horizontal="right")
ws1.row_dimensions[r].height = 20; r += 1

# ════════════════════════════════════════════════════════════════
#  SHEET 2 — BALANCE SHEET
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Balance Sheet")
col_widths(ws2, [28, 4, 4, 18])

r = 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws2.cell(row=r, column=1, value=data["sacco_name"].upper())
c.font = font(bold=True, color=C_DARK, size=13)
c.alignment = Alignment(horizontal="center")
ws2.row_dimensions[r].height = 22; r += 1

ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws2.cell(row=r, column=1, value="BALANCE SHEET")
c.font = font(bold=True, color=C_TEAL, size=11)
c.alignment = Alignment(horizontal="center")
ws2.row_dimensions[r].height = 18; r += 1

ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws2.cell(row=r, column=1, value=f"As at {data['end_date']}")
c.font = font(italic=True, color=C_MUTED, size=9)
c.alignment = Alignment(horizontal="center")
ws2.row_dimensions[r].height = 14; r += 1

blank_row(ws2, r); r += 1

for col, lbl in [(1,"Description"), (4,"Amount (UGX)")]:
    c = ws2.cell(row=r, column=col, value=lbl)
    c.font = font(bold=True, color=C_WHITE)
    c.fill = hdr_fill(C_DARK)
    c.alignment = Alignment(horizontal="left" if col==1 else "right", indent=1)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws2.row_dimensions[r].height = 16; r += 1

# ASSETS
section_header(ws2, r, "ASSETS", C_TEAL, C_WHITE); r += 1
ast_start = r
for i, item in enumerate(data["asset_items"]):
    data_row(ws2, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
ast_end = r - 1
total_row(ws2, r, "TOTAL ASSETS", f"=SUM(D{ast_start}:D{ast_end})", bg="E6FAF4", fg=C_DARK)
total_assets_row = r; r += 1

blank_row(ws2, r); r += 1

# LIABILITIES
section_header(ws2, r, "LIABILITIES", C_RED, C_WHITE); r += 1
lib_start = r
for i, item in enumerate(data["liability_items"]):
    data_row(ws2, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
lib_end = r - 1
total_row(ws2, r, "TOTAL LIABILITIES", f"=SUM(D{lib_start}:D{lib_end})", bg="FFF5F5", fg=C_RED)
total_lib_row = r; r += 1

blank_row(ws2, r); r += 1

# EQUITY / FUNDS
section_header(ws2, r, "MEMBERS' FUNDS & EQUITY", "2D3748", C_WHITE); r += 1
eq_start = r
for i, item in enumerate(data["equity_items"]):
    data_row(ws2, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
eq_end = r - 1
total_row(ws2, r, "TOTAL MEMBERS' FUNDS", f"=SUM(D{eq_start}:D{eq_end})", bg="EBF8FF", fg=C_DARK)
total_eq_row = r; r += 1

blank_row(ws2, r); r += 1

# CHECK: Liabilities + Equity = Assets
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = ws2.cell(row=r, column=1, value="TOTAL LIABILITIES + MEMBERS' FUNDS")
c.font = font(bold=True, color=C_WHITE, size=11)
c.fill = hdr_fill(C_DARK)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
v = ws2.cell(row=r, column=4, value=f"=D{total_lib_row}+D{total_eq_row}")
v.font = font(bold=True, color=C_WHITE, size=11)
v.number_format = UGX
v.fill = hdr_fill(C_DARK)
v.alignment = Alignment(horizontal="right")
ws2.row_dimensions[r].height = 20; r += 1

# ════════════════════════════════════════════════════════════════
#  SHEET 3 — LOAN PORTFOLIO
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Loan Portfolio")
col_widths(ws3, [28, 4, 4, 18])

r = 1
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws3.cell(row=r, column=1, value=data["sacco_name"].upper())
c.font = font(bold=True, color=C_DARK, size=13)
c.alignment = Alignment(horizontal="center")
ws3.row_dimensions[r].height = 22; r += 1

ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws3.cell(row=r, column=1, value="LOAN PORTFOLIO SUMMARY")
c.font = font(bold=True, color=C_TEAL, size=11)
c.alignment = Alignment(horizontal="center")
ws3.row_dimensions[r].height = 18; r += 1

ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws3.cell(row=r, column=1, value=f"As at {data['end_date']}")
c.font = font(italic=True, color=C_MUTED, size=9)
c.alignment = Alignment(horizontal="center")
ws3.row_dimensions[r].height = 14; r += 1

blank_row(ws3, r); r += 1

for col, lbl in [(1,"Description"), (4,"Amount (UGX)")]:
    c = ws3.cell(row=r, column=col, value=lbl)
    c.font = font(bold=True, color=C_WHITE)
    c.fill = hdr_fill(C_DARK)
    c.alignment = Alignment(horizontal="left" if col==1 else "right", indent=1)
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws3.row_dimensions[r].height = 16; r += 1

section_header(ws3, r, "LOAN PORTFOLIO", C_TEAL, C_WHITE); r += 1
lp_start = r
for i, item in enumerate(data["loan_items"]):
    data_row(ws3, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1
lp_end = r - 1

blank_row(ws3, r); r += 1
section_header(ws3, r, "INTEREST ANALYSIS", "2D3748", C_WHITE); r += 1
ia_start = r
for i, item in enumerate(data["interest_items"]):
    data_row(ws3, r, item["label"], item["amount"], zebra=(i%2==0)); r += 1

wb.save("/mnt/user-data/outputs/financial_statements.xlsx")
print("OK")
